import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font

# Solicita o input do usuário para o termo de busca
from selenium.webdriver.support.wait import WebDriverWait

termo_busca = input("Digite a Vaga ou empresa que deseja buscar: ")

# Configura o driver do selenium ou o driver do navegador
options = webdriver.ChromeOptions()
# options.add_argument("--headless")
# options.add_argument("--no-sandbox")
# options.add_argument("--disable-dev-shm-usage")
driver = webdriver.Chrome(options=options)
driver.get(f"https://portal.gupy.io/job-search/term={termo_busca}")

# Encontra o elemento HTML que contém as informações de emprego
vagas = driver.find_elements('xpath','//*[@id="__next"]/div[3]/div/div/main/ul')

# Define o tempo total para rolagem da página em segundos
scroll_time = 30
start_time = time.time()

# Rola a página para baixo e espera carregar
while True:
    # Verifica se o tempo total para rolagem foi alcançado
    if time.time() - start_time > scroll_time:
        break

    driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
    time.sleep(5)
    vagas = driver.find_elements('xpath', '//*[@id="__next"]/div[3]/div/div/main/ul/div/li/div')

    # Aguarda até que o scroll esteja ocioso antes de tentar encontrar novos elementos
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'li')))
    vagas = driver.find_elements('xpath', '//*[@id="__next"]/div[3]/div/div/main/ul/div/li/div')

    # Verifica se não há mais vagas para exibir
    final_page = driver.find_elements('xpath', '//*[@id="__next"]/div[3]/div/div/main/ul/div')
    if final_page == True:
        break
    else:
        continue

# Analisa o conteúdo HTML da página com a biblioteca BeautifulSoup
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
# Formata o HTML com o método prettify()
# html_formatted = soup.prettify()
# print(html_formatted)

jobs = soup.find('ul', {'class': 'sc-90466136-0 djVYjM'})

# Cria uma nova planilha Excel para armazenar as informações coletadas
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Vagas de emprego"
header = ['Titulo da Vaga', 'Empresa', 'Localidade', 'Tipo Contratação', 'Links da Vaga', 'Data Extração']
sheet.append(header)
# Configurando Detalhes da Planilha
bold_font = Font(bold=True) # define a fonte em negrito
# define o estilo de fonte em negrito para cada célula do header
for cell in sheet[1]:
    cell.font = bold_font

# Retornando a data/Hora atual
now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

# Itera sobre as vagas encontradas e armazena as informações relevantes na planilha Excel
linha = 2
for job in jobs.find_all('li'):
    titulo_vaga = job.find('h4').text.strip()
    empresa = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-5 hBmurm'}).text.strip()
    localizacao = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 sc-812f417a-6 dkBRQd kaqRPR'}).text.strip()
    tipo_contratacao = job.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 dkBRQd'}).text.strip()
    links = [link['href'] for link in job.find_all('a', {'class': 'sc-812f417a-1 fijAgW'})]
    time.sleep(5)
    print(f"\n"
          f"Vaga: {titulo_vaga}\n"
          f"Empresa: {empresa}\n"
          f"Local: {localizacao}\n"
          f"Tipo Contrato: {tipo_contratacao}\n"
          f"Links das Vagas: {links}\n"
          f"Data de Coleta e Envio dos Dados: {now}")

    sheet.cell(row=linha, column=1).value = titulo_vaga
    sheet.cell(row=linha, column=2).value = empresa
    sheet.cell(row=linha, column=3).value = localizacao
    sheet.cell(row=linha, column=4).value = tipo_contratacao
    sheet.cell(row=linha, column=5).value = ", ".join(links)
    sheet.cell(row=linha, column=6).value = now

    linha += 1
    time.sleep(1)
# Salva a planilha Excel
workbook.save(f"{termo_busca}_vagas.xlsx")

# Fecha o driver do selenium
driver.quit()

print(f"As vagas de emprego para '{termo_busca}' foram salvas em uma planilha Excel.")
