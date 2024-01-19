#RETORNA AS 10 PRIMEIRAS VAGAS

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font
import sys
import time

# Faz a solicitação GET para a página de busca com o termo inserido pelo usuário
url_base = "https://portal.gupy.io/job-search/"
offset = 0
vagas = []

headers = {'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                        '(KHTML, like Gecko) Chrome/111.0.0.0 Safari/537.36'}

# Solicita o input do usuário para o termo de busca
termo_busca = input("Digite a Vaga ou Empresa que deseja buscar: ")
remoto = input("Deseja apenas vagas remoto? (sim ou não) ")
if remoto.lower() == "sim":
    remoto = "only-remote"
else:
    remoto = ""

pcd = input("Deseja vagas para PCD? (sim ou não) ")
if pcd.lower() == "sim":
    pcd = "true"
else:
    pcd = ""

if remoto == 'sim' and pcd == 'sim':
    url = f"https://portal.gupy.io/job-search/term={termo_busca}&remoteWorking={remoto}&pwd={pcd}"
elif remoto == 'sim' and pcd == 'não':
    url = f"https://portal.gupy.io/job-search/term={termo_busca}&remoteWorking={remoto}"
elif remoto == 'não' and pcd == 'sim':
    url = f"https://portal.gupy.io/job-search/term={termo_busca}&pwd={pcd}"
else:
    url = f"https://portal.gupy.io/job-search/term={termo_busca}"

print(url)

# Faz a solicitação GET para a página de busca com o termo inserido pelo usuário
url = f"https://portal.gupy.io/job-search/term={termo_busca}"
response = requests.get(url, headers=headers)
time.sleep(3)

# Analisa o conteúdo HTML da página com a biblioteca BeautifulSoup
soup = BeautifulSoup(response.content, 'html.parser')

# Encontra o elemento HTML que contém as informações de emprego
vagas = soup.find('ul', {'class': 'sc-90466136-0 djVYjM'})
if vagas:
    print("Tudo certo, continue!")
else:
    vagas = soup.find('h1', {'class': 'sc-llJcti hVUvys sc-68db55b3-0 fqSWza'}).text.strip()
    print(f"{vagas}, tente outras funções, Obrigado!")
    sys.exit()


# Cria uma nova planilha Excel para armazenar as informações coletadas
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Vagas de emprego"
header = ['Titulo da Vaga', 'Empresa', 'Localidade', 'Tipo Contratação', 'Links da Vaga']
sheet.append(header)
# Configurando Detalhes da Planilha
bold_font = Font(bold=True) # define a fonte em negrito
# define o estilo de fonte em negrito para cada célula do header
for cell in sheet[1]:
    cell.font = bold_font

# Itera sobre as vagas encontradas e armazena as informações relevantes na planilha Excel
linha = 2
for vaga in vagas.find_all('li'):
    titulo_vaga = vaga.find('h4').text.strip()
    empresa = vaga.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-5 hBmurm'}).text.strip()
    localizacao = vaga.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 sc-812f417a-6 dkBRQd kaqRPR'}).text.strip()
    tipo_contratacao = vaga.find('p', {'class': 'sc-efBctP dpAAMR sc-812f417a-4 dkBRQd'}).text.strip()
    links = [link['href'] for link in vaga.find_all('a', {'class': 'sc-812f417a-1 fijAgW'})]
    time.sleep(15)
    print(f"\n"
          f"Vaga: {titulo_vaga}\n"
          f"Empresa: {empresa}\n"
          f"Local: {localizacao}\n"
          f"Tipo Contrato: {tipo_contratacao}\n"
          f"Links das Vagas: {links}")

    sheet.cell(row=linha, column=1).value = titulo_vaga
    sheet.cell(row=linha, column=2).value = empresa
    sheet.cell(row=linha, column=3).value = localizacao
    sheet.cell(row=linha, column=4).value = tipo_contratacao
    sheet.cell(row=linha, column=5).value = ", ".join(links)

    linha += 1

# Salva a planilha Excel
workbook.save(f"{termo_busca}_10_primeira_vagas.xlsx")

print(f"As vagas de emprego para '{termo_busca}' foram salvas em uma planilha Excel.")
